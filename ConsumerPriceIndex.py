# Part B: Consumer Price Index
# Client wants to analyze CPI data for Canada and each province. 
# Client wants to use this new class to read multiple CPI files (one for each jurisdiction), combine them into a single data frame, and perform a series of inflation and cost-of-living calculations.
# Client wants to use this library to:
#   - Show first 12 rows of data.
#   - Calculate average monthly % change for Food, Shelter, and All-items (excl. food & energy).
#   - Identify the province with the highest change.
#   - Use All-items CPI to find equivalent salaries across provinces (e.g., $100K in Ontario, Dec-24).
#   - From MinimumWages.csv, find highest/lowest nominal and highest real (CPI-adjusted) wages.
#   - Compute annual % change in Services CPI and identify the region with the highest inflation.
# Attributes of the CPIAnalysis class (Part B version) are:
#   - cpi_files: a list or dictionary of file paths for the 11 CPI CSV files
#   - wages_file: the file path for MinimumWages.csv
# Public methods:
#   - load_cpi_data(): read and combine all CPI files.
#   - print_head(): display first 12 rows of combined data.
#   - avg_monthly_changes(): calculate average monthly % change for selected CPI items.
#   - province_with_highest_change(): find province with the highest average change.
#   - equivalent_salary(): compute equivalent salaries across provinces using All-items CPI.
#   - min_wage_summary(): find highest/lowest nominal and highest real (CPI-adjusted) wages.
#   - services_inflation(): calculate annual % change in Services CPI.
#   - region_with_highest_services_inflation(): identify region with highest services inflation.
# Assumes values have been validated (file paths, column names, months, and items are correct and consistent with the Statistics Canada CPI data and MinimumWages.csv).

import pandas as pd

"""Edit to match the full path to your CSV folder which contains unzipped folders"""
filepath = r"C:\Users\Owner\OneDrive - York University\Documents\FINE 3300\A2 Data\A2 Data"

csv_files = [
    filepath + "\\Canada.CPI.1810000401.csv",
    filepath + "\\AB.CPI.1810000401.csv",
    filepath + "\\BC.CPI.1810000401.csv",
    filepath + "\\MB.CPI.1810000401.csv",
    filepath + "\\NB.CPI.1810000401.csv",
    filepath + "\\NL.CPI.1810000401.csv",
    filepath + "\\NS.CPI.1810000401.csv",
    filepath + "\\ON.CPI.1810000401.csv",
    filepath + "\\PEI.CPI.1810000401.csv",
    filepath + "\\QC.CPI.1810000401.csv",
    filepath + "\\SK.CPI.1810000401.csv"
]

# Question 1
df = pd.DataFrame()

for file_name in csv_files:
    temp = pd.melt(pd.read_csv(file_name), id_vars="Item", var_name="Month", value_name="CPI")

    """Clean jurisdiction name from the filename (e.g., 'ON', 'BC', 'Canada')"""
    basename = file_name.replace("\\", "/").split("/")[-1]
    temp["Jurisdiction"] = basename.split(".")[0]

    temp["Month"] = temp["Month"].apply(
        lambda x: x.replace("24-", "") + "-24" if x.startswith("24-") else x
    )

    df = pd.concat([df, temp], ignore_index=True)

# Question 2
"""Client wants to display the first 12 rows of the combined CPI data frame"""
pd.set_option('display.max_colwidth', None)  # show full text in columns

df = df[["Item", "Month", "Jurisdiction", "CPI"]]
print("QUESTION 2: FIRST 12 LINES OF THE COMBINED DATAFRAME:")
print(df.head(12))

# Convert Month to a proper datetime once
df["Date"] = pd.to_datetime(df["Month"], format="%b-%y")

# Convert Month to a proper datetime once
df["Date"] = pd.to_datetime(df["Month"], format="%b-%y")

# Question 3
"""Client wants to calculate the average month-to-month % change for Food, Shelter, and All-items (excluding food & energy),and display them in a wide table by province and Canada."""
df["CPI"] = pd.to_numeric(df["CPI"])
df["MoM_pct_change"] = df.groupby(["Jurisdiction", "Item"])["CPI"].pct_change() * 100
"""Select only items of interest"""
items_of_interest = ["All-items excluding food and energy", "Food", "Shelter"]
df_interest = df.loc[df["Item"].isin(items_of_interest)]
"""Compute average monthly change by jurisdiction and item"""
avg_change = (
    df_interest
    .groupby(["Jurisdiction", "Item"])["MoM_pct_change"]
    .mean()
    .reset_index()
)
"""Pivot to make items columns (wide format)"""
avg_pivot = avg_change.pivot(index="Jurisdiction", columns="Item", values="MoM_pct_change")
"""# Round to one decimal and format with '%'"""
avg_pivot = avg_pivot.round(1).astype(str) + "%"
"""Sort alphabetically by jurisdiction"""
avg_pivot = avg_pivot.sort_index()
print("\nQUESTION 3: AVERAGE MONTH-TO-MONTH CHANGE (DISPLAYED BY CATEGORY):")
print(avg_pivot.reset_index())

# Question 4
"""Client wants to identify all provinces that have the highest average change in each category."""
"""Work from the numeric summary created in Question 3 (avg_change) and exclude Canada from the comparison"""
prov_only = avg_change.loc[avg_change["Jurisdiction"] != "Canada"].copy()

"""Round once to avoid tiny floating-point differences"""
prov_only["MoM_rounded"] = prov_only["MoM_pct_change"].round(1)

"""For each Item, find the maximum rounded value"""
max_by_item = prov_only.groupby("Item")["MoM_rounded"].transform("max")

"""Keep all provinces whose rounded value equals the max for that Item (handles ties)"""
top_prov = prov_only.loc[prov_only["MoM_rounded"] == max_by_item].copy()

"""Sort for nicer display"""
top_prov = top_prov.sort_values(["Item", "Jurisdiction"]).reset_index(drop=True)

print("\nQUESTION 4: PROVINCE(S) WITH THE HIGHEST AVERAGE CHANGE FOR EACH ITEM:")
for item, group in top_prov.groupby("Item"):
    prov_list = ", ".join(group["Jurisdiction"])
    change_val = group["MoM_rounded"].iloc[0]
    print(f"{item}: {prov_list} - {change_val}%")


# Question 5
"""Client wants to compute equivalent salaries across provinces using Dec-24 All-items CPI and display both CPI and equivalent salary values."""
all_items = df[df["Item"] == "All-items"]

"""Filter for December 2024"""
dec24 = all_items[
    (all_items["Date"].dt.year == 2024) &
    (all_items["Date"].dt.month == 12)
].copy()

"""Ontario’s CPI used as base reference"""
on_cpi = dec24.loc[dec24["Jurisdiction"] == "ON", "CPI"].iloc[0]

"""Compute equivalent salary in each province based on CPI ratio"""
dec24["Equivalent_Salary"] = (100000 * dec24["CPI"] / on_cpi).round(2)

"""Format CPI and salary nicely"""
dec24["CPI"] = dec24["CPI"].round(1)
dec24["Equivalent_Salary"] = dec24["Equivalent_Salary"].apply(lambda x: f"${x:,.2f}")

"""Keep only columns we want to show"""
equiv_salaries = dec24.loc[
    ~dec24["Jurisdiction"].isin(["ON", "Canada"]),
    ["Jurisdiction", "CPI", "Equivalent_Salary"]
].reset_index(drop=True)

print("\nQUESTION 5: EQUIVALENT SALARY TO $100,000 RECEIVED IN ONTARIO IN ALL OTHER PROVINCES (AS AT DECEMBER 2024):")
print(equiv_salaries)

# Question 6
"""Read minimum wage data"""
min_wages = pd.read_csv(filepath + r"\MinimumWages.csv")

"""Find highest and lowest nominal minimum wages"""
highest_nominal = min_wages.loc[min_wages["Minimum Wage"].idxmax()]
lowest_nominal = min_wages.loc[min_wages["Minimum Wage"].idxmin()]

print("\nQUESTION 6: HIGHEST AND LOWEST MINIMUM WAGE ON NOMINAL BASIS:")
print(f"Highest Minimum Wage - {highest_nominal['Province']}: ${highest_nominal['Minimum Wage']:.2f}")
print(f"Lowest Minimum Wage - {lowest_nominal['Province']}: ${lowest_nominal['Minimum Wage']:.2f}")

"""Use CPI data for December 2024"""
all_items = df[df["Item"] == "All-items"]
dec24 = all_items[all_items["Month"] == "Dec-24"].copy()

"""Rename CPI column to match Province column for merge and compute Real Minimum Wage using (Nominal / CPI) × 100"""
dec24 = dec24.rename(columns={"Jurisdiction": "Province"})
merged = pd.merge(min_wages, dec24[["Province", "CPI"]], on="Province", how="inner")

merged["Real_Min_Wage"] = (merged["Minimum Wage"] / merged["CPI"]) * 100
merged["Nominal_Real_Diff"] = merged["Minimum Wage"] - merged["Real_Min_Wage"]

"""Find province with highest real minimum wage"""
highest_real = merged.loc[merged["Real_Min_Wage"].idxmax()]

print("\nQUESTION 6: HIGHEST REAL MINIMUM WAGE USING CPI NUMBERS FOR DECEMBER 2024:")
print(f"{highest_real['Province']}: ${highest_real['Real_Min_Wage']:.2f}")

"""Display CPI, nominal, and real wage difference by province"""
print("\nQUESTION 6: NOMINAL VS REAL MINIMUM WAGE DIFFERENCE (DECEMBER 2024):")
diff_table = merged[["Province", "CPI", "Minimum Wage", "Real_Min_Wage", "Nominal_Real_Diff"]].copy()
diff_table["CPI"] = diff_table["CPI"].round(1)
diff_table["Minimum Wage"] = diff_table["Minimum Wage"].round(2)
diff_table["Real_Min_Wage"] = diff_table["Real_Min_Wage"].round(2)
diff_table["Nominal_Real_Diff"] = diff_table["Nominal_Real_Diff"].round(2)
"""Add $ signs for wage columns"""
diff_table["Minimum Wage"] = diff_table["Minimum Wage"].apply(lambda x: f"${x:,.2f}")
diff_table["Real_Min_Wage"] = diff_table["Real_Min_Wage"].apply(lambda x: f"${x:,.2f}")
diff_table["Nominal_Real_Diff"] = diff_table["Nominal_Real_Diff"].apply(lambda x: f"${x:,.2f}")

print(diff_table.to_string(index=False))

# Question 7
"""Client wants to calculate the annual % change in CPI for Services across all jurisdictions"""
"""Filter only rows where Item = "Services" since CPI includes many categories and group by province/region and get the first and last CPI readings for the year"""
df_services = df.loc[df["Item"] == "Services"]
annual_stats = df_services.groupby("Jurisdiction")["CPI"].agg(["first", "last"]).reset_index()
annual_stats["Annual_pct_change"] = ((annual_stats["last"] - annual_stats["first"]) / annual_stats["first"]) * 100
annual_stats["Annual_pct_change"] = annual_stats["Annual_pct_change"].round(1)

print("\nQUESTION 7: ANNUAL CHANGE IN CPI FOR SERVICES (FIRST VS. LAST MONTH):")
display_stats = annual_stats.copy()
display_stats["Annual_pct_change"] = display_stats["Annual_pct_change"].map(lambda x: f"{x:.1f}%")
print(display_stats)

top_services = annual_stats.loc[annual_stats["Annual_pct_change"].idxmax()]

# Question 8
"""Client wants to identify the region with the highest annual inflation in Services"""
print("\nQUESTION 8: REGION WITH THE HIGHEST INFLATION IN SERVICES:")
print("Jurisdiction:", top_services["Jurisdiction"])
print("Inflation Value (%):", f"{top_services['Annual_pct_change']:.1f}%")

# SAVE ALL QUESTION OUTPUTS TO ONE EXCEL FILE
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

excel_out = "CPI_Analysis_Results.xlsx"

with pd.ExcelWriter(excel_out, engine="openpyxl") as writer:
    # Q2 – first 12 rows
    df[["Item", "Month", "Jurisdiction", "CPI"]].head(12).to_excel(
        writer, sheet_name="Q2_First12Rows", index=False
    )

    # Q3 – average month-to-month % change pivot table
    avg_pivot.reset_index().to_excel(writer, sheet_name="Q3_AvgMoMChange", index=False)

    # Q4 – provinces with highest average change
    top_prov.to_excel(writer, sheet_name="Q4_TopProvince", index=False)

    # Q5 – equivalent salary table
    equiv_salaries.to_excel(writer, sheet_name="Q5_EquivalentSalary", index=False)

    # Q6 – nominal vs real minimum wage difference
    diff_table.to_excel(writer, sheet_name="Q6_NominalVsRealDiff", index=False)


    # Q7 – services inflation table
    annual_stats.to_excel(writer, sheet_name="Q7_ServicesInflation", index=False)

    # Q8 – single-row summary of top region
    pd.DataFrame([top_services]).to_excel(writer, sheet_name="Q8_TopRegion", index=False)

    # Apply formatting to every worksheet 
    wb = writer.book
    for ws in wb.worksheets:
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for col_idx, col_name in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            col_letter = get_column_letter(col_idx)
            header_cell = ws[f"{col_letter}1"]
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_cell.fill = header_fill

            # Auto-fit column width
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0 for cell in col_name),
                default=len(col_name),
            )
            ws.column_dimensions[col_letter].width = max_len + 2

        # Apply currency format only to Q6’s wage columns
        if ws.title == "Q6_NominalVsRealDiff":
            for col in ["C", "D", "E"]:  # Minimum Wage, Real_Min_Wage, Nominal_Real_Diff
                for cell in ws[col][1:]:
                    cell.number_format = u'"$"#,##0.00'

        # Apply percent-like format to Annual_pct_change for Q7 and Q8
        if ws.title in ("Q7_ServicesInflation", "Q8_TopRegion"):
            # Find the column that contains 'Annual_pct_change'
            for cell in ws[1]:
                if str(cell.value) == "Annual_pct_change":
                    pct_col_letter = cell.column_letter
                    # Format as number with a literal % sign (no scaling)
                    for c in ws[pct_col_letter][1:]:
                        c.number_format = '0.0"%"'
                    break

print(f"\nAll question outputs have been saved to: {excel_out}")
