# Part A: Loan Amortization and Payment Schedule
# Client wants to expand the MortgagePayment class from Assignment 1 to also prompt the user for the term of the mortgage alognside already prompting for the principal, annual quoted interest rate (%), compounded semi-annually, and amortization period (years). 
# Client wants to use this library to calculate mortgage calculations for 6 payment options: monthly, semi-monthly, bi-weekly, weekly, rapid bi-weekly, rapid weekly. 
# Attributes of the MortgagePayment class (extended version): rate (interest rate in decimal form), amort_years (amortization period in years), term_years (mortgage term in years used for schedule length)
# Public Method: payments(principal)
#                takes the loan principal amount as a parameter
#                returns a tuple of payment amounts for the 6 frequencies using the present value of an annuity factor formula.
#                build a loan payment schedule (amortization table) using pandas
#                generates dataFrames including the columns: Period, Starting Balance, Interest, Payment, Ending Balance
#                save all 6 schedules into one Excel file with 6 worksheets (one worksheet per payment option)
#                plot the loan balance decline for all 6 schedules on the same chart
# Assumes values have been validated. 

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from matplotlib.ticker import FuncFormatter 


class MortgagePayment:
    """Mortgage calculations in Canada are computed as follows: 
    The interest rate is provided as a quoted semi-annual rate which is convered to an Effective Annual Rate (EAR)
    The loan's amortization is the amount of time over which the loan's payment is calculated. 
    The length of the payment schedule is based on the term of the mortgage."""

    def __init__(self, rate_percent, amort_years, term_years):
        self.rate = rate_percent / 100.0     # convert % to decimal
        self.amort_years = int(amort_years)
        self.term_years = int(term_years)

        # Convert the nominal semi-annual rate to an effective annual rate.
        # EAR = (1 + r/2)^2 - 1
        self.ear = (1 + self.rate / 2.0) ** 2 - 1

    def _periodic_rate(self, freq_per_year):
        """Convert the effective annual rate to per-period rate."""
        return (1 + self.ear) ** (1 / freq_per_year) - 1

    def _num_payments(self, years, freq_per_year):
        """Compute the number of payments for a given horizon and frequency."""
        return int(years * freq_per_year)

    def _annuity_factor(self, r, n):
        """Compute the present value of an annuity-immediate factor."""
        if r == 0:
            return n
        return (1 - (1 + r) ** (-n)) / r

    def _payment_amount(self, principal, r, n):
        """Compute the constant level payment using the annuity factor."""
        return principal / self._annuity_factor(r, n)

    def payments(self, principal):
        """Calculate payment amounts for 6 frequencies and return as tuple."""
        freqs = (12, 24, 26, 52)
        base_payments = []

        # Compute regular payments
        for f in freqs:
            r = self._periodic_rate(f)
            n = self._num_payments(self.amort_years, f)
            pay = self._payment_amount(principal, r, n)
            base_payments.append(round(pay, 2))

        # Rapid options based on monthly payment
        monthly = base_payments[0]
        base_payments += [round(monthly / 2, 2), round(monthly / 4, 2)]

        return tuple(base_payments)

    def payment_details(self, principal):
        """Build a dictionary with payment details for all six options."""
        freq_map = {
            "Monthly": 12,
            "Semi-Monthly": 24,
            "Bi-Weekly": 26,
            "Weekly": 52,
        }

        details = {}
        for name, f in freq_map.items():
            r = self._periodic_rate(f)
            n_amort = self._num_payments(self.amort_years, f)
            n_term = self._num_payments(self.term_years, f)
            pay = self._payment_amount(principal, r, n_amort)
            details[name] = (pay, r, n_term)

        # Rapid plans
        monthly_pay, _, _ = details["Monthly"]
        details["Rapid Bi-Weekly"] = (monthly_pay / 2, self._periodic_rate(26), self._num_payments(self.term_years, 26))
        details["Rapid Weekly"] = (monthly_pay / 4, self._periodic_rate(52), self._num_payments(self.term_years, 52))
        return details

    def _make_schedule(self, principal, payment, r, n_term):
        """Build a single payment schedule DataFrame."""
        periods, start_bal, interests, pay_list, end_bal = [], [], [], [], []
        balance = float(principal)

        for t in range(1, n_term + 1):
            if balance <= 1e-8:
                break
            interest = balance * r
            this_payment = min(payment, balance + interest)
            new_balance = balance + interest - this_payment

            periods.append(t)
            start_bal.append(balance)
            interests.append(interest)
            pay_list.append(this_payment)
            end_bal.append(new_balance)
            balance = new_balance

        df = pd.DataFrame({
            "Period": periods,
            "Starting Balance": np.round(start_bal, 2),
            "Interest Amount": np.round(interests, 2),
            "Payment": np.round(pay_list, 2),
            "Ending Balance": np.round(end_bal, 2)
        })
        return df

    def build_all_schedules(self, principal):
        """Generate all 6 amortization schedules."""
        info = self.payment_details(principal)
        return {name: self._make_schedule(principal, pay, r, n_term) for name, (pay, r, n_term) in info.items()}


if __name__ == "__main__":
    # Prompt user for input values
    print("Mortgage Payment Calculator")
    print("---------------------------")
    principal = float(input("Enter the principal amount ($): "))
    rate = float(input("Enter the annual quoted interest rate (%): "))
    amort_years = int(input("Enter the amortization period (years): "))
    term_years = int(input("Enter the term of the mortgage (years): "))

    mortgage = MortgagePayment(rate, amort_years, term_years)

    print("\nMortgage Payment Options")
    print("------------------------")
    details = mortgage.payment_details(principal)
    for name, (pay, _, _) in details.items():
        print(f"{name:>18}: ${pay:,.2f}")

    # Build the six DataFrame schedules
    schedules = mortgage.build_all_schedules(principal)

    # Save to Excel with sheet names ending in "Payments"
    excel_file = "Payment_Schedules.xlsx"
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        for name, df in schedules.items():
            sheet_name = f"{name} Payments"
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            # Header formatting
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                header_cell = ws[f"{col_letter}1"]
                header_cell.font = Font(bold=True)
                header_cell.alignment = Alignment(horizontal="center")
                header_cell.fill = header_fill

                # Auto-fit column width
                max_len = max((len(str(cell.value)) for cell in ws[col_letter]), default=len(col_name))
                ws.column_dimensions[col_letter].width = max_len + 2

                # Format numeric columns as currency
                if col_name != "Period":
                    for cell in ws[col_letter][1:]:
                        cell.number_format = u'"$"#,##0.00'

    print(f"\nAll 6 mortgage schedules have been saved to: {excel_file}")

  # Plot balance decline for all 6 schedules
    plt.figure(figsize=(10, 7))
    for name, df in schedules.items():
        plt.plot(df["Period"], df["Ending Balance"], label=name)
    plt.title("Loan Balance Decline by Payment Schedule")
    plt.xlabel("Payment Period (within term)")
    plt.ylabel("Ending Balance ($)")
    plt.legend()
    plt.grid(True, linestyle="--", linewidth=0.5, alpha=0.6)

    # Format Y-axis with commas (e.g., 100,000 instead of 100000)
    formatter = FuncFormatter(lambda x, _: f"{x:,.0f}")
    plt.gca().yaxis.set_major_formatter(formatter)

    plt.tight_layout()

    png_file = "Loan_Balance_Decline.png"
    plt.savefig(png_file, dpi=200)
    plt.close()

    print(f"Loan balance decline graph saved to: {png_file}")